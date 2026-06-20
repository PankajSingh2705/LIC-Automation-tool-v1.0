# ============================================
# LIC AUTOMATION v1.0
# File : writer.py
# ============================================

import os
from openpyxl import Workbook, load_workbook


class ExcelWriter:

    HEADERS = [
        "Policy Number",
        "Name",
        "DOC",
        "Plan Name",
        "Term",
        "PPT",
        "Premium",
        "Payment Mode",
        "Next Premium",
        "Policy Status",
        "Maturity Date"
    ]

    FAILED_HEADERS = [
        "Policy Number",
        "Reason"
    ]

    def __init__(self, output_file, failed_file):

        self.output_file = output_file
        self.failed_file = failed_file

        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        self._create_output_file()
        self._create_failed_file()

    # ------------------------------------

    def _create_output_file(self):

        if os.path.exists(self.output_file):
            return

        wb = Workbook()
        ws = wb.active

        ws.append(self.HEADERS)

        wb.save(self.output_file)

    # ------------------------------------

    def _create_failed_file(self):

        if os.path.exists(self.failed_file):
            return

        wb = Workbook()
        ws = wb.active

        ws.append(self.FAILED_HEADERS)

        wb.save(self.failed_file)

    # ------------------------------------

    def save_policy(self, data):

        wb = load_workbook(self.output_file)
        ws = wb.active

        ws.append([
            data.get("Policy Number", ""),
            data.get("Name", ""),
            data.get("DOC", ""),
            data.get("Plan Name", ""),
            data.get("Term", ""),
            data.get("PPT", ""),
            data.get("Premium", ""),
            data.get("Payment Mode", ""),
            data.get("Next Premium", ""),
            data.get("Policy Status", ""),
            data.get("Maturity Date", "")
        ])

        wb.save(self.output_file)

    # ------------------------------------

    def save_failed(self, policy, reason):

        wb = load_workbook(self.failed_file)
        ws = wb.active

        ws.append([
            policy,
            reason
        ])

        wb.save(self.failed_file)

    # ------------------------------------

    def total_success(self):

        wb = load_workbook(self.output_file)
        ws = wb.active

        return ws.max_row - 1

    # ------------------------------------

    def total_failed(self):

        wb = load_workbook(self.failed_file)
        ws = wb.active

        return ws.max_row - 1