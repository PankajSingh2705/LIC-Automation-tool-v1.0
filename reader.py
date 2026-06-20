# ============================================
# LIC AUTOMATION v1.0
# File : reader.py
# ============================================

from openpyxl import load_workbook
import os


class PolicyReader:

    def __init__(self, excel_file):
        self.excel_file = excel_file

        if not os.path.exists(excel_file):
            raise FileNotFoundError(
                f"\nInput Excel not found:\n{excel_file}"
            )

        self.workbook = load_workbook(excel_file)
        self.sheet = self.workbook.active

    # ----------------------------------------

    def get_total_policies(self):
        """
        Total rows excluding header
        """
        return self.sheet.max_row - 1

    # ----------------------------------------

    def read_all_policies(self):

        policies = []

        for row in range(2, self.sheet.max_row + 1):

            value = self.sheet.cell(row=row, column=1).value

            if value is None:
                continue

            value = str(value).strip()

            if value == "":
                continue

            policies.append(value)

        return policies

    # ----------------------------------------

    def print_summary(self):

        policies = self.read_all_policies()

        print("\n==============================")
        print("INPUT EXCEL SUMMARY")
        print("==============================")
        print(f"Total Policies : {len(policies)}")
        print("==============================\n")

    # ----------------------------------------

    def get_policy(self, index):

        policies = self.read_all_policies()

        if index >= len(policies):
            return None

        return policies[index]